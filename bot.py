import os
import re
import html
import shutil
import logging
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pdfplumber
from mega import Mega
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfReader, PdfWriter

from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.constants import ChatAction
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

# =========================================================
# НАСТРОЙКИ ИЗ ПЕРЕМЕННЫХ ОКРУЖЕНИЯ
# =========================================================
TOKEN = os.getenv("BOT_TOKEN", "").strip()
MEGA_EMAIL = os.getenv("MEGA_EMAIL", "").strip()
MEGA_PASSWORD = os.getenv("MEGA_PASSWORD", "").strip()
ADMIN_IDS_RAW = os.getenv("ADMIN_IDS", "").strip()

ALLOWED_USERS = {
    int(x.strip()) for x in ADMIN_IDS_RAW.split(",") if x.strip().isdigit()
}

# Папки на Mega
MEGA_ROOT = "151-4"
MEGA_ORIGINAL = f"{MEGA_ROOT}/Original"
MEGA_KVIT = f"{MEGA_ROOT}/Kvitancii"
MEGA_CLIENTS = f"{MEGA_ROOT}/Klienty"

DB_FILENAME = "База данных.xlsx"
DB_HEADERS = ["Телефон", "Рахунок", "Ф.І.О."]

# =========================================================
# СОСТОЯНИЯ ДИАЛОГОВ
# =========================================================
ADD_PHONE, ADD_ACCOUNT, ADD_FIO = range(3)

DEL_PHONE_OR_SKIP, DEL_ACCOUNT_OR_SKIP, DEL_PICK = range(10, 13)

EDIT_PHONE_OR_SKIP, EDIT_ACCOUNT_OR_SKIP, EDIT_PICK, EDIT_NEW_PHONE, EDIT_NEW_ACCOUNT, EDIT_NEW_FIO = range(20, 26)

logger = logging.getLogger(__name__)


# =========================================================
# КЛАВИАТУРЫ
# =========================================================
def main_menu() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton("Завантажити та Розділити")],
            [KeyboardButton("Додати рахунок")],
            [KeyboardButton("Видалити рахунок")],
            [KeyboardButton("Редагувати рахунок")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def cancel_button() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[KeyboardButton("Скасувати")]],
        resize_keyboard=True,
        is_persistent=True,
    )


# =========================================================
# ДОСТУП ПО ADMIN_IDS
# =========================================================
def is_admin(update: Update) -> bool:
    if not ALLOWED_USERS:
        return True
    user = update.effective_user
    return bool(user and user.id in ALLOWED_USERS)


async def require_admin(update: Update) -> bool:
    if is_admin(update):
        return True

    if update.message:
        await update.message.reply_text(
            "У вас немає доступу до цього бота.",
            reply_markup=ReplyKeyboardRemove(),
        )
    elif update.callback_query:
        await update.callback_query.answer("Немає доступу", show_alert=True)
    return False


# =========================================================
# ВСПОМОГАТЕЛЬНОЕ
# =========================================================
def normalize_phone(value: str) -> str:
    return re.sub(r"\D+", "", value or "")


def validate_phone(value: str) -> bool:
    return value.isdigit() and len(value) == 12


def normalize_account(value: str) -> str:
    return re.sub(r"\s+", "", (value or "").strip())


def sanitize_filename(name: str) -> str:
    name = re.sub(r'[\\/*?:"<>|]+', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:180] or "file"


def extract_fields_from_text(text: str) -> Tuple[str, str, str]:
    """
    Пытается вытащить:
    - год
    - месяц
    - номер особового рахунку

    Если твой PDF имеет другой шаблон текста, чаще всего надо будет
    подправить ТОЛЬКО regex в этой функции.
    """
    text = text or ""

    account_patterns = [
        r"Особов(?:ий|ого)\s+рахунок[:\s№#]*([0-9]{5,20})",
        r"Особовий\s*р\/?х[:\s№#]*([0-9]{5,20})",
        r"Лицев(?:ой|ого)\s+счет[:\s№#]*([0-9]{5,20})",
        r"\bрахунок[:\s№#]*([0-9]{5,20})",
    ]

    year_patterns = [
        r"\b(20\d{2})\b",
    ]

    month_name_map = {
        "січень": "01",
        "лютий": "02",
        "березень": "03",
        "квітень": "04",
        "травень": "05",
        "червень": "06",
        "липень": "07",
        "серпень": "08",
        "вересень": "09",
        "жовтень": "10",
        "листопад": "11",
        "грудень": "12",
        "январь": "01",
        "февраль": "02",
        "март": "03",
        "апрель": "04",
        "май": "05",
        "июнь": "06",
        "июль": "07",
        "август": "08",
        "сентябрь": "09",
        "октябрь": "10",
        "ноябрь": "11",
        "декабрь": "12",
    }

    account = None
    for pattern in account_patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            account = m.group(1)
            break

    if not account:
        all_numbers = re.findall(r"\b\d{6,20}\b", text)
        account = all_numbers[0] if all_numbers else "UNKNOWN"

    year = None
    for pattern in year_patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            year = m.group(1)
            break
    if not year:
        year = "0000"

    month = None

    # Формат 03.2026 / 3-2026 / 03/2026
    m = re.search(r"\b(0?[1-9]|1[0-2])[./ -](20\d{2})\b", text)
    if m:
        month = m.group(1).zfill(2)
        year = m.group(2)

    # Формат 2026-03 / 2026/03
    if not month:
        m = re.search(r"\b(20\d{2})[./ -](0?[1-9]|1[0-2])\b", text)
        if m:
            year = m.group(1)
            month = m.group(2).zfill(2)

    # Місяць: 3 / 03
    if not month:
        m = re.search(r"(?:Місяць|Месяц|Період|Период)[:\s]*([0-1]?\d)", text, re.IGNORECASE)
        if m:
            month = m.group(1).zfill(2)

    # Название месяца словами
    if not month:
        lowered = text.lower()
        for name, num in month_name_map.items():
            if re.search(rf"\b{name}\b", lowered):
                month = num
                break

    if not month:
        month = "00"

    return year, month, account


# =========================================================
# MEGA-ХРАНИЛИЩЕ
# =========================================================
class MegaStorage:
    def __init__(self, email: str, password: str):
        self.email = email
        self.password = password
        self.client = None

    def connect(self):
        if self.client is None:
            mega = Mega()
            self.client = mega.login(self.email, self.password)
        return self.client

    def ensure_folder(self, path: str):
        m = self.connect()
        parts = [p for p in path.strip("/").split("/") if p]

        current_path = ""
        current_node = None

        for part in parts:
            current_path = f"{current_path}/{part}" if current_path else part
            found = m.find(current_path)

            if not found:
                m.create_folder(current_path)
                found = m.find(current_path)

            if not found:
                raise RuntimeError(f"Не вдалося знайти або створити папку: {current_path}")

            current_node = found[0] if isinstance(found, list) else found

        return current_node

    def find(self, path: str):
        m = self.connect()
        return m.find(path)

    def delete_if_exists(self, path: str):
        m = self.connect()
        found = m.find(path)
        if not found:
            return

        if isinstance(found, list):
            for item in found:
                try:
                    m.delete(item)
                except Exception:
                    pass
        else:
            try:
                m.delete(found)
            except Exception:
                pass

       def upload_file(self, local_path: str, mega_folder_path: str, mega_name: Optional[str] = None):
        m = self.connect()

        try:
            m.create_folder(mega_folder_path)
        except Exception:
            pass

        if mega_name:
            temp_path = Path(local_path).with_name(mega_name)
            shutil.copy2(local_path, temp_path)
            try:
                return m.upload(str(temp_path), dest=mega_folder_path)
            finally:
                if temp_path.exists():
                    temp_path.unlink(missing_ok=True)

        return m.upload(local_path, dest=mega_folder_path)
    def download_file(self, mega_file_path: str, local_dir: str) -> Optional[str]:
        m = self.connect()
        file_node = m.find(mega_file_path)
        if not file_node:
            return None
        downloaded = m.download(file_node, local_dir)
        return downloaded


storage = MegaStorage(MEGA_EMAIL, MEGA_PASSWORD)


# =========================================================
# EXCEL
# =========================================================
def ensure_local_workbook(path: str):
    if Path(path).exists():
        wb = load_workbook(path)
        ws = wb.active

        # Если файл существует, но пустой
        if ws.max_row == 1 and all(ws.cell(1, i + 1).value is None for i in range(3)):
            ws.append(DB_HEADERS)
            wb.save(path)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Клієнти"
    ws.append(DB_HEADERS)
    wb.save(path)


def fetch_database_to_temp(tmp_dir: str) -> str:
    storage.ensure_folder(MEGA_CLIENTS)
    db_mega_path = f"{MEGA_CLIENTS}/{DB_FILENAME}"
    local_path = os.path.join(tmp_dir, DB_FILENAME)

    downloaded = storage.download_file(db_mega_path, tmp_dir)
    if downloaded:
        if downloaded != local_path and Path(downloaded).exists():
            shutil.move(downloaded, local_path)

    ensure_local_workbook(local_path)
    return local_path


def save_database_back(local_path: str):
    db_mega_path = f"{MEGA_CLIENTS}/{DB_FILENAME}"
    storage.delete_if_exists(db_mega_path)
    storage.upload_file(local_path, MEGA_CLIENTS, DB_FILENAME)


def read_rows(local_db_path: str) -> List[Dict]:
    wb = load_workbook(local_db_path)
    ws = wb.active

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        phone = ws[f"A{row_idx}"].value
        acc = ws[f"B{row_idx}"].value
        fio = ws[f"C{row_idx}"].value

        if phone is None and acc is None and fio is None:
            continue

        rows.append(
            {
                "row": row_idx,
                "phone": str(phone or "").strip(),
                "account": str(acc or "").strip(),
                "fio": str(fio or "").strip(),
            }
        )
    return rows


def append_row(local_db_path: str, phone: str, account: str, fio: str):
    wb = load_workbook(local_db_path)
    ws = wb.active
    ws.append([phone, account, fio])
    wb.save(local_db_path)


def delete_row(local_db_path: str, row_number: int):
    wb = load_workbook(local_db_path)
    ws = wb.active
    ws.delete_rows(row_number, 1)
    wb.save(local_db_path)


def update_row(local_db_path: str, row_number: int, phone: str, account: str, fio: str):
    wb = load_workbook(local_db_path)
    ws = wb.active
    ws[f"A{row_number}"] = phone
    ws[f"B{row_number}"] = account
    ws[f"C{row_number}"] = fio
    wb.save(local_db_path)


def find_records(local_db_path: str, phone: str, account: str) -> List[Dict]:
    phone = normalize_phone(phone)
    account = normalize_account(account)

    results = []
    for row in read_rows(local_db_path):
        row_phone = normalize_phone(row["phone"])
        row_account = normalize_account(row["account"])

        phone_ok = (not phone) or (row_phone == phone)
        acc_ok = (not account) or (row_account == account)

        # Если заполнены оба поля — выводим по совпадению одного ИЛИ второго,
        # как в ТЗ "по телефону или по рахунку".
        if phone and account:
            if row_phone == phone or row_account == account:
                results.append(row)
        else:
            if phone_ok and acc_ok:
                results.append(row)

    return results


def build_pick_keyboard(action: str, records: List[Dict]) -> InlineKeyboardMarkup:
    buttons = []
    for rec in records:
        text = f'{action}: {rec["account"]} | {rec["phone"]}'
        buttons.append([InlineKeyboardButton(text, callback_data=f"{action}|{rec['row']}")])
    return InlineKeyboardMarkup(buttons)


# =========================================================
# START / CANCEL
# =========================================================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update):
        return
    await update.message.reply_text("Оберіть дію.", reply_markup=main_menu())


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update):
        return ConversationHandler.END

    context.chat_data.clear()
    await update.message.reply_text("Скасовано.", reply_markup=main_menu())
    return ConversationHandler.END


# =========================================================
# МЕНЮ
# =========================================================
async def menu_click(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update):
        return

    text = (update.message.text or "").strip()

    if text == "Завантажити та Розділити":
        context.chat_data["await_pdf"] = True
        await update.message.reply_text(
            "Надішліть PDF-файл одним повідомленням.\n"
            "Я завантажу оригінал у 151-4/Original, розділю по сторінках і кожну сторінку завантажу у 151-4/Kvitancii.",
            reply_markup=cancel_button(),
        )
        return

    if text == "Додати рахунок":
        await update.message.reply_text(
            "Введіть номер телефону з 12 цифр без +.\n"
            "Наприклад: 380XXXXXXXXX",
            reply_markup=cancel_button(),
        )
        return ADD_PHONE

    if text == "Видалити рахунок":
        await update.message.reply_text(
            "Введіть телефон з 12 цифр без + або надішліть - щоб пропустити цей крок.",
            reply_markup=cancel_button(),
        )
        return DEL_PHONE_OR_SKIP

    if text == "Редагувати рахунок":
        await update.message.reply_text(
            "Введіть телефон з 12 цифр без + або надішліть - щоб пропустити цей крок.",
            reply_markup=cancel_button(),
        )
        return EDIT_PHONE_OR_SKIP

    await update.message.reply_text("Оберіть кнопку з меню.", reply_markup=main_menu())


# =========================================================
# PDF
# =========================================================
async def handle_pdf_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update):
        return

    if not context.chat_data.get("await_pdf"):
        return

    document = update.message.document
    if not document:
        await update.message.reply_text("Чекаю саме PDF-файл.")
        return

    if not document.file_name.lower().endswith(".pdf"):
        await update.message.reply_text("Це не PDF. Надішліть файл з розширенням .pdf")
        return

    context.chat_data["await_pdf"] = False
    await update.message.chat.send_action(ChatAction.UPLOAD_DOCUMENT)

    with tempfile.TemporaryDirectory() as tmp_dir:
        original_name = sanitize_filename(document.file_name)
        local_pdf = os.path.join(tmp_dir, original_name)

        tg_file = await document.get_file()
        await tg_file.download_to_drive(local_pdf)

        # Убедимся, что папки существуют
        storage.ensure_folder(MEGA_ORIGINAL)
        storage.ensure_folder(MEGA_KVIT)

        # Загружаем оригинал
        storage.upload_file(local_pdf, MEGA_ORIGINAL, original_name)

        reader = PdfReader(local_pdf)
        uploaded_count = 0

        with pdfplumber.open(local_pdf) as pdf:
            for index, page in enumerate(reader.pages):
                writer = PdfWriter()
                writer.add_page(page)

                page_text = ""
                try:
                    page_text = pdf.pages[index].extract_text() or ""
                except Exception:
                    page_text = ""

                year, month, account = extract_fields_from_text(page_text)
                split_name = sanitize_filename(f"{year}_{month}_{account}.pdf")
                split_path = os.path.join(tmp_dir, split_name)

                with open(split_path, "wb") as out:
                    writer.write(out)

                storage.upload_file(split_path, MEGA_KVIT, split_name)
                uploaded_count += 1

    await update.message.reply_text(
        f"Готово. Оригінальний PDF завантажено в «Original», а {uploaded_count} окремих файлів — у «Kvitancii».",
        reply_markup=main_menu(),
    )


async def unknown_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update):
        return
    await update.message.reply_text(
        "Спочатку натисніть «Завантажити та Розділити», а потім надішліть PDF.",
        reply_markup=main_menu(),
    )


# =========================================================
# ДОБАВИТЬ РАХУНОК
# =========================================================
async def add_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    value = normalize_phone(update.message.text)

    if not validate_phone(value):
        await update.message.reply_text(
            "Телефон має містити рівно 12 цифр без +. Спробуйте ще раз."
        )
        return ADD_PHONE

    context.chat_data["new_phone"] = value
    await update.message.reply_text("Введіть номер особового рахунку.")
    return ADD_ACCOUNT


async def add_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    value = normalize_account(update.message.text)

    if not value:
        await update.message.reply_text("Рахунок не може бути порожнім. Введіть номер рахунку.")
        return ADD_ACCOUNT

    context.chat_data["new_account"] = value
    await update.message.reply_text("Введіть Ф.І.О. або надішліть - якщо поле порожнє.")
    return ADD_FIO


async def add_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    fio = update.message.text.strip()
    if fio == "-":
        fio = ""

    phone = context.chat_data["new_phone"]
    account = context.chat_data["new_account"]

    with tempfile.TemporaryDirectory() as tmp_dir:
        local_db = fetch_database_to_temp(tmp_dir)
        append_row(local_db, phone, account, fio)
        save_database_back(local_db)

    context.chat_data.clear()

    await update.message.reply_text(
        f"Рахунок {html.escape(account)} додано.",
        reply_markup=main_menu(),
    )
    return ConversationHandler.END


# =========================================================
# УДАЛИТЬ РАХУНОК
# =========================================================
async def delete_phone_or_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()

    if raw == "-":
        context.chat_data["search_phone"] = ""
        await update.message.reply_text("Введіть рахунок або надішліть - щоб шукати тільки за телефоном.")
        return DEL_ACCOUNT_OR_SKIP

    value = normalize_phone(raw)
    if not validate_phone(value):
        await update.message.reply_text(
            "Телефон має містити рівно 12 цифр без + або надішліть -"
        )
        return DEL_PHONE_OR_SKIP

    context.chat_data["search_phone"] = value
    await update.message.reply_text("Введіть рахунок або надішліть - щоб шукати тільки за телефоном.")
    return DEL_ACCOUNT_OR_SKIP


async def delete_account_or_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    account = "" if raw == "-" else normalize_account(raw)
    phone = context.chat_data.get("search_phone", "")

    if not phone and not account:
        await update.message.reply_text("Треба вказати хоча б телефон або рахунок.")
        return DEL_ACCOUNT_OR_SKIP

    with tempfile.TemporaryDirectory() as tmp_dir:
        local_db = fetch_database_to_temp(tmp_dir)
        found = find_records(local_db, phone, account)

    if not found:
        context.chat_data.clear()
        await update.message.reply_text("Нічого не знайдено.", reply_markup=main_menu())
        return ConversationHandler.END

    context.chat_data["delete_found"] = found

    await update.message.reply_text(
        "Знайдено записи. Натисніть, що саме видалити.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await update.message.reply_text(
        "Виберіть рядок для видалення:",
        reply_markup=build_pick_keyboard("delete", found),
    )
    return DEL_PICK


async def delete_pick_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update):
        return ConversationHandler.END

    query = update.callback_query
    await query.answer()

    action, row_str = query.data.split("|", 1)
    if action != "delete":
        return DEL_PICK

    row_number = int(row_str)
    account_removed = ""

    with tempfile.TemporaryDirectory() as tmp_dir:
        local_db = fetch_database_to_temp(tmp_dir)
        rows = read_rows(local_db)
        selected = next((r for r in rows if r["row"] == row_number), None)

        if not selected:
            await query.edit_message_text("Запис не знайдено або вже видалено.")
            context.chat_data.clear()
            await query.message.reply_text("Повертаюсь у меню.", reply_markup=main_menu())
            return ConversationHandler.END

        account_removed = selected["account"]
        delete_row(local_db, row_number)
        save_database_back(local_db)

    context.chat_data.clear()
    await query.edit_message_text(f"Рахунок {account_removed} видалено.")
    await query.message.reply_text("Оберіть наступну дію.", reply_markup=main_menu())
    return ConversationHandler.END


# =========================================================
# РЕДАКТИРОВАТЬ РАХУНОК
# =========================================================
async def edit_phone_or_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()

    if raw == "-":
        context.chat_data["search_phone"] = ""
        await update.message.reply_text("Введіть рахунок або надішліть - щоб шукати тільки за телефоном.")
        return EDIT_ACCOUNT_OR_SKIP

    value = normalize_phone(raw)
    if not validate_phone(value):
        await update.message.reply_text(
            "Телефон має містити рівно 12 цифр без + або надішліть -"
        )
        return EDIT_PHONE_OR_SKIP

    context.chat_data["search_phone"] = value
    await update.message.reply_text("Введіть рахунок або надішліть - щоб шукати тільки за телефоном.")
    return EDIT_ACCOUNT_OR_SKIP


async def edit_account_or_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    account = "" if raw == "-" else normalize_account(raw)
    phone = context.chat_data.get("search_phone", "")

    if not phone and not account:
        await update.message.reply_text("Треба вказати хоча б телефон або рахунок.")
        return EDIT_ACCOUNT_OR_SKIP

    with tempfile.TemporaryDirectory() as tmp_dir:
        local_db = fetch_database_to_temp(tmp_dir)
        found = find_records(local_db, phone, account)

    if not found:
        context.chat_data.clear()
        await update.message.reply_text("Нічого не знайдено.", reply_markup=main_menu())
        return ConversationHandler.END

    context.chat_data["edit_found"] = found

    await update.message.reply_text(
        "Знайдено записи. Натисніть, який саме редагувати.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await update.message.reply_text(
        "Виберіть рядок для редагування:",
        reply_markup=build_pick_keyboard("edit", found),
    )
    return EDIT_PICK


async def edit_pick_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update):
        return ConversationHandler.END

    query = update.callback_query
    await query.answer()

    action, row_str = query.data.split("|", 1)
    if action != "edit":
        return EDIT_PICK

    row_number = int(row_str)
    found = context.chat_data.get("edit_found", [])
    selected = next((r for r in found if r["row"] == row_number), None)

    if not selected:
        await query.edit_message_text("Запис не знайдено.")
        await query.message.reply_text("Повертаюсь у меню.", reply_markup=main_menu())
        context.chat_data.clear()
        return ConversationHandler.END

    context.chat_data["edit_row"] = row_number
    context.chat_data["edit_old_account"] = selected["account"]

    await query.edit_message_text(
        "Редагування запису.\n"
        f"Поточний телефон: {selected['phone']}\n"
        f"Поточний рахунок: {selected['account']}\n"
        f"Поточне Ф.І.О.: {selected['fio'] or '—'}"
    )

    await query.message.reply_text("Введіть новий телефон з 12 цифр без +.")
    return EDIT_NEW_PHONE


async def edit_new_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    value = normalize_phone(update.message.text)

    if not validate_phone(value):
        await update.message.reply_text(
            "Телефон має містити рівно 12 цифр без +. Спробуйте ще раз."
        )
        return EDIT_NEW_PHONE

    context.chat_data["edit_phone"] = value
    await update.message.reply_text("Введіть новий рахунок.")
    return EDIT_NEW_ACCOUNT


async def edit_new_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    value = normalize_account(update.message.text)

    if not value:
        await update.message.reply_text("Рахунок не може бути порожнім. Введіть новий рахунок.")
        return EDIT_NEW_ACCOUNT

    context.chat_data["edit_account"] = value
    await update.message.reply_text("Введіть нове Ф.І.О. або надішліть - якщо треба залишити порожнім.")
    return EDIT_NEW_FIO


async def edit_new_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    fio = update.message.text.strip()
    if fio == "-":
        fio = ""

    row_number = context.chat_data["edit_row"]
    phone = context.chat_data["edit_phone"]
    account = context.chat_data["edit_account"]

    with tempfile.TemporaryDirectory() as tmp_dir:
        local_db = fetch_database_to_temp(tmp_dir)
        update_row(local_db, row_number, phone, account, fio)
        save_database_back(local_db)

    old_account = context.chat_data.get("edit_old_account", account)
    context.chat_data.clear()

    await update.message.reply_text(
        f"Рахунок {old_account} відредаговано.",
        reply_markup=main_menu(),
    )
    return ConversationHandler.END


# =========================================================
# ОБРАБОТКА ОШИБОК
# =========================================================
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.exception("Unhandled error", exc_info=context.error)

    if isinstance(update, Update) and update.effective_message:
        try:
            await update.effective_message.reply_text("Сталася помилка під час роботи з MEGA або базою даних.")
        except Exception:
            pass


# =========================================================
# СБОРКА ПРИЛОЖЕНИЯ
# =========================================================
def build_application() -> Application:
    if not TOKEN:
        raise RuntimeError("Заповніть BOT_TOKEN у змінних середовища.")

    app = Application.builder().token(TOKEN).build()

    add_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Додати рахунок$"), menu_click)],
        states={
            ADD_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_phone)],
            ADD_ACCOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_account)],
            ADD_FIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_fio)],
        },
        fallbacks=[MessageHandler(filters.Regex("^Скасувати$"), cancel)],
        allow_reentry=True,
    )

    del_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Видалити рахунок$"), menu_click)],
        states={
            DEL_PHONE_OR_SKIP: [MessageHandler(filters.TEXT & ~filters.COMMAND, delete_phone_or_skip)],
            DEL_ACCOUNT_OR_SKIP: [MessageHandler(filters.TEXT & ~filters.COMMAND, delete_account_or_skip)],
            DEL_PICK: [CallbackQueryHandler(delete_pick_callback, pattern=r"^delete\|")],
        },
        fallbacks=[MessageHandler(filters.Regex("^Скасувати$"), cancel)],
        allow_reentry=True,
    )

    edit_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Редагувати рахунок$"), menu_click)],
        states={
            EDIT_PHONE_OR_SKIP: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_phone_or_skip)],
            EDIT_ACCOUNT_OR_SKIP: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_account_or_skip)],
            EDIT_PICK: [CallbackQueryHandler(edit_pick_callback, pattern=r"^edit\|")],
            EDIT_NEW_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_new_phone)],
            EDIT_NEW_ACCOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_new_account)],
            EDIT_NEW_FIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_new_fio)],
        },
        fallbacks=[MessageHandler(filters.Regex("^Скасувати$"), cancel)],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(add_conv)
    app.add_handler(del_conv)
    app.add_handler(edit_conv)

    app.add_handler(MessageHandler(filters.Regex("^Завантажити та Розділити$"), menu_click))
    app.add_handler(MessageHandler(filters.Document.PDF, handle_pdf_document))
    app.add_handler(MessageHandler(filters.Document.ALL, unknown_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, menu_click))

    app.add_error_handler(error_handler)
    return app


if __name__ == "__main__":
    logging.basicConfig(
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
        level=logging.INFO,
    )

    application = build_application()
    application.run_polling(allowed_updates=Update.ALL_TYPES)
